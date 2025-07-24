import streamlit as st
import pandas as pd
import pdfplumber
import re
import os
from datetime import datetime
import io
from openpyxl import Workbook, load_workbook
import plotly.express as px
import plotly.graph_objects as go

# ページ設定
st.set_page_config(page_title="地方銀行財務データ抽出", layout="wide")

# データベース管理関数
def initialize_database(db_path="data/securities_database.xlsx"):
    """データベースを初期化する"""
    # dataディレクトリが存在しない場合は作成
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    if not os.path.exists(db_path):
        # 新規データベースを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "証券データ"
        
        # ヘッダーを設定
        headers = ['年月', '国債', '地方債', '短期社債', '社債', '株式', '外国証券', 'その他の証券', '貸出金', '更新日時']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(db_path)
        st.info(f"新しいデータベースを作成しました: {db_path}")
    
    return db_path

def save_to_database(data, db_path="data/securities_database.xlsx"):
    """データをExcelデータベースに保存する"""
    try:
        # データベースを初期化
        initialize_database(db_path)
        
        # Excelファイルを読み込み
        wb = load_workbook(db_path)
        ws = wb.active
        
        # 現在の日時を追加
        data_with_timestamp = data.copy()
        data_with_timestamp['更新日時'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 既存データから同じ年月のレコードを探す
        existing_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == data['年月']:
                existing_row = row
                break
        
        # データを書き込む列の順序
        columns = ['年月', '国債', '地方債', '短期社債', '社債', '株式', '外国証券', 'その他の証券', '貸出金', '更新日時']
        
        if existing_row:
            # 既存レコードを更新
            for col, column_name in enumerate(columns, 1):
                ws.cell(row=existing_row, column=col, value=data_with_timestamp[column_name])
            action = "updated"
        else:
            # 新規レコードを追加
            new_row = ws.max_row + 1
            for col, column_name in enumerate(columns, 1):
                ws.cell(row=new_row, column=col, value=data_with_timestamp[column_name])
            action = "added"
        
        # データを年月順でソートしてから保存
        # まずデータをDataFrameに変換
        data_for_sorting = []
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, len(columns) + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            if any(cell is not None for cell in row_data):  # 空行でない場合のみ追加
                data_for_sorting.append(row_data)
        
        # DataFrameを作成してソート
        if data_for_sorting:
            df_sort = pd.DataFrame(data_for_sorting, columns=columns)
            # 年月列でソート（昇順：古いデータから新しいデータへ）
            df_sort = df_sort.sort_values('年月', ascending=True)
            
            # ワークシートをクリア（ヘッダー以外）
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
            
            # ソートされたデータを書き戻し
            for idx, (_, row_data) in enumerate(df_sort.iterrows(), start=2):
                for col, column_name in enumerate(columns, 1):
                    ws.cell(row=idx, column=col, value=row_data[column_name])
        
        # Excelファイルを保存
        wb.save(db_path)
        
        return True, action
    
    except Exception as e:
        return False, str(e)

def load_database(db_path="data/securities_database.xlsx"):
    """データベースからデータを読み込む"""
    try:
        if os.path.exists(db_path):
            df = pd.read_excel(db_path)
            return df
        else:
            return pd.DataFrame()
    except Exception as e:
        st.error(f"データベースの読み込みエラー: {str(e)}")
        return pd.DataFrame()

def extract_securities_from_pdf(pdf_file):
    """PDFから証券データを抽出する"""
    try:
        # pdfplumberでPDFを開く
        with pdfplumber.open(pdf_file) as pdf:
            # 2ページ目を取得（0ベースなので1）
            if len(pdf.pages) < 2:
                st.error("PDFファイルに2ページ目が見つかりません")
                return None
                
            page = pdf.pages[1]
            
            # テキストを抽出
            page_text = page.extract_text()
            
            # デバッグ用：抽出したテキストを表示
            with st.expander("🔍 抽出されたPDFテキスト（デバッグ用）"):
                st.text_area("PDFテキスト", page_text, height=400)
            
            # テーブルを抽出
            tables = page.extract_tables()
            
            # デバッグ用：テーブル構造を表示
            with st.expander("🔍 抽出されたテーブル構造（デバッグ用）"):
                for i, table in enumerate(tables):
                    st.write(f"**テーブル {i+1}:**")
                    if table:
                        try:
                            # 列名の重複を処理
                            headers = table[0] if table[0] else None
                            if headers:
                                # 重複した列名に番号を追加
                                unique_headers = []
                                header_counts = {}
                                for header in headers:
                                    if header in header_counts:
                                        header_counts[header] += 1
                                        unique_headers.append(f"{header}_{header_counts[header]}")
                                    else:
                                        header_counts[header] = 1
                                        unique_headers.append(header)
                                
                                df = pd.DataFrame(table[1:], columns=unique_headers)
                            else:
                                df = pd.DataFrame(table)
                            
                            st.dataframe(df)
                        except Exception as e:
                            st.write(f"テーブル表示エラー: {str(e)}")
                            # エラーの場合は生データを表示
                            st.write("生データ:")
                            for j, row in enumerate(table[:5]):  # 最初の5行のみ表示
                                st.write(f"行{j}: {row}")
            
            # 年月の抽出（例：2025年3月中平残）
            year_month_match = re.search(r'(\d{4})年\s*(\d{1,2})月', page_text)
            if year_month_match:
                year = year_month_match.group(1)
                month = year_month_match.group(2).zfill(2)
                year_month = f"{year}-{month}"
            else:
                year_month = "不明"
            
            # 証券の金額を抽出
            def extract_securities_amounts(text, tables):
                """各証券の金額を抽出する"""
                debug_info = []
                
                # 抽出対象の証券項目（スペースを含む形式も考慮）
                # 注意: 検索順序が重要！より具体的なパターンを先に検索する
                securities_patterns = {
                    '貸出金': ['貸 出 金', '貸出金'],  # 貸出金を最初に追加
                    '短期社債': ['短 期 社 債', '短期社債'],  # 「社債」より先に検索
                    '社債': ['社 債', '社債'],                # 「短期社債」の後に検索
                    '国債': ['国 債', '国債'],
                    '地方債': ['地 方 債', '地方債'],
                    '株式': ['株 式', '株式'],
                    '外国証券': ['外 国 証 券', '外国証券'],
                    'その他の証券': ['そ の 他 の 証 券', 'その他の証券']
                }
                
                securities_data = {}
                
                # 各証券項目を初期化
                for security in securities_patterns.keys():
                    securities_data[security] = "0"
                
                # 方法1: テキストから直接抽出（スペースを考慮）
                debug_info.append("テキストから直接抽出を開始...")
                
                # テキストを行に分割
                lines = text.split('\n')
                
                for security, patterns in securities_patterns.items():
                    debug_info.append(f"{security}を検索中...")
                    
                    for pattern in patterns:
                        for line_idx, line in enumerate(lines):
                            # 社債の場合は短期社債と区別するための特別処理
                            if security == '社債' and pattern in line:
                                # 「短期」が含まれていないことを確認
                                if '短期' not in line and '短 期' not in line:
                                    debug_info.append(f"{security}を含む行を発見 (パターン: '{pattern}', 行{line_idx}): {line}")
                                    
                                    # パターンの後にある数値を抽出
                                    pattern_index = line.find(pattern)
                                    if pattern_index >= 0:
                                        # パターンの後の部分を取得
                                        after_pattern = line[pattern_index + len(pattern):]
                                        
                                        # 最初の数値を探す
                                        numbers = re.findall(r'([0-9,]+)', after_pattern)
                                        for num in numbers:
                                            clean_num = num.replace(',', '')
                                            if clean_num.isdigit() and len(clean_num) >= 4:
                                                debug_info.append(f"{security}の金額を発見: {clean_num}")
                                                securities_data[security] = clean_num
                                                break
                                        
                                        if securities_data[security] != "0":
                                            break
                                else:
                                    debug_info.append(f"{security}の候補行をスキップ (短期社債と判定): {line}")
                            
                            elif security != '社債' and pattern in line:
                                debug_info.append(f"{security}を含む行を発見 (パターン: '{pattern}', 行{line_idx}): {line}")
                                
                                # パターンの後にある数値を抽出
                                pattern_index = line.find(pattern)
                                if pattern_index >= 0:
                                    # パターンの後の部分を取得
                                    after_pattern = line[pattern_index + len(pattern):]
                                    
                                    # 最初の数値を探す
                                    numbers = re.findall(r'([0-9,]+)', after_pattern)
                                    for num in numbers:
                                        clean_num = num.replace(',', '')
                                        if clean_num.isdigit() and len(clean_num) >= 4:
                                            debug_info.append(f"{security}の金額を発見: {clean_num}")
                                            securities_data[security] = clean_num
                                            break
                                    
                                    if securities_data[security] != "0":
                                        break
                        
                        # 見つかったら他のパターンを試さない
                        if securities_data[security] != "0":
                            break
                
                # 方法2: 正規表現による抽出（スペースを考慮した柔軟なパターン）
                debug_info.append("正規表現による抽出を開始...")
                
                for security, patterns in securities_patterns.items():
                    if securities_data[security] == "0":
                        debug_info.append(f"{security}の正規表現マッチングを試行...")
                        
                        for pattern in patterns:
                            # 社債の場合は短期社債を除外する正規表現を使用
                            if security == '社債':
                                # 負の先読みを使用して「短期」が前にない「社債」を抽出
                                flexible_pattern = pattern.replace(' ', r'\s*')
                                regex_patterns = [
                                    rf'(?<!短\s*期\s*){flexible_pattern}\s+([0-9,]+)',
                                    rf'(?<!短期){re.escape(pattern)}\s+([0-9,]+)',
                                ]
                            else:
                                # 通常の処理
                                flexible_pattern = pattern.replace(' ', r'\s*')
                                regex_patterns = [
                                    rf'{flexible_pattern}\s+([0-9,]+)',
                                    rf'{re.escape(pattern)}\s+([0-9,]+)',
                                ]
                            
                            for regex_pattern in regex_patterns:
                                match = re.search(regex_pattern, text)
                                if match:
                                    found_amount = match.group(1).replace(',', '')
                                    if found_amount.isdigit() and len(found_amount) >= 4:
                                        debug_info.append(f"{security}を正規表現で抽出 (パターン: '{pattern}'): {found_amount}")
                                        securities_data[security] = found_amount
                                        break
                            
                            if securities_data[security] != "0":
                                break
                
                # 方法3: テーブルから抽出（最後の手段）
                debug_info.append("テーブルからの抽出を開始...")
                
                for table_idx, table in enumerate(tables):
                    if not table:
                        continue
                    
                    debug_info.append(f"テーブル {table_idx + 1} を処理中...")
                    
                    for row_idx, row in enumerate(table):
                        if not row:
                            continue
                        
                        row_text = ' '.join([str(cell) if cell else '' for cell in row])
                        
                        for security, patterns in securities_patterns.items():
                            if securities_data[security] == "0":
                                for pattern in patterns:
                                    if pattern in row_text:
                                        debug_info.append(f"テーブルで{security}を含む行を発見 (テーブル{table_idx+1}, 行{row_idx}): {row}")
                                        
                                        # この行から数値を抽出
                                        for cell_idx, cell in enumerate(row):
                                            if cell and isinstance(cell, str):
                                                numbers = re.findall(r'([0-9,]+)', cell)
                                                for num in numbers:
                                                    clean_num = num.replace(',', '')
                                                    if clean_num.isdigit() and len(clean_num) >= 4:
                                                        debug_info.append(f"テーブルで{security}の金額を発見: {clean_num}")
                                                        securities_data[security] = clean_num
                                                        break
                                                if securities_data[security] != "0":
                                                    break
                                        if securities_data[security] != "0":
                                            break
                                if securities_data[security] != "0":
                                    break
                
                return securities_data, debug_info
            
            # 証券金額を抽出
            securities_data, debug_info = extract_securities_amounts(page_text, tables)
            
            # デバッグ情報を表示
            with st.expander("🔍 抽出結果の詳細（デバッグ用）"):
                st.write("**抽出処理のログ:**")
                for info in debug_info:
                    st.write(f"- {info}")
                
                # 証券キーワードを含む行をテーブルから探す
                st.write("**証券を含むテーブル行の詳細:**")
                securities_patterns = {
                    '貸出金': ['貸 出 金', '貸出金', '貸 出'],  # 貸出金を最初に追加
                    '短期社債': ['短 期 社 債', '短期社債'],  # 「社債」より先に検索
                    '社債': ['社 債', '社債'],                # 「短期社債」の後に検索
                    '国債': ['国 債', '国債'],
                    '地方債': ['地 方 債', '地方債'],
                    '株式': ['株 式', '株式'],
                    '外国証券': ['外 国 証 券', '外国証券'],
                    'その他の証券': ['そ の 他 の 証 券', 'その他の証券', 'そ の 他 証 券']
                }
                
                for i, table in enumerate(tables):
                    if not table:
                        continue
                    st.write(f"テーブル {i+1}:")
                    for j, row in enumerate(table):
                        if row:
                            row_text = ' '.join([str(cell) if cell else '' for cell in row])
                            for security, patterns in securities_patterns.items():
                                for pattern in patterns:
                                    if pattern in row_text:
                                        st.write(f"  {security} (パターン: '{pattern}') - 行{j}: {row}")
                                        # 各セルの数値をチェック
                                        for k, cell in enumerate(row):
                                            if cell:
                                                numbers = re.findall(r'([0-9,]+)', str(cell))
                                                if numbers:
                                                    st.write(f"    セル{k} '{cell}' から数値: {numbers}")
                                        break
                                if any(pattern in row_text for pattern in patterns):
                                    break
                
                st.write("**最終抽出結果:**")
                for security, amount in securities_data.items():
                    st.write(f"- {security}: {amount}")
            
            return {
                '年月': year_month,
                **{k: v for k, v in securities_data.items() if k != '貸出金'},  # 貸出金以外を先に追加
                '貸出金': securities_data.get('貸出金', '0'),  # 貸出金を最後に追加
            }
        
    except Exception as e:
        st.error(f"PDFの読み込み中にエラーが発生しました: {str(e)}")
        return None

# メインアプリケーション
st.title("🏦 地方銀行財務データ抽出システム")
st.markdown("---")

# サイドバー
with st.sidebar:
    st.header("ページ選択")
    page = st.selectbox(
        "表示するページを選択してください",
        ["データ抽出", "グラフ表示"]
    )
    
    st.markdown("---")
    
    if page == "データ抽出":
        st.header("ファイルアップロード")
        uploaded_file = st.file_uploader(
            "PDFファイルを選択してください",
            type=['pdf'],
            help="地方銀行主要勘定のPDFファイルをアップロードしてください"
        )
    else:
        uploaded_file = None

# メインエリア
if page == "データ抽出":
    # データ抽出ページ
    if uploaded_file is not None:
        # PDFからデータを抽出
        with st.spinner("PDFから証券データを抽出中..."):
            extracted_data = extract_securities_from_pdf(uploaded_file)
        
        if extracted_data:
            st.success("✅ 証券データの抽出が完了しました")
            
            # 抽出したデータを表示
            st.subheader("📊 抽出結果")
            
            # 年月を表示
            st.write(f"**年月: {extracted_data['年月']}**")
            
            # 証券データを3列で表示
            col1, col2, col3 = st.columns(3)
            
            securities_list = ['国債', '地方債', '短期社債', '社債', '株式', '外国証券', 'その他の証券', '貸出金']
            
            for i, security in enumerate(securities_list):
                col_idx = i % 3
                amount = int(extracted_data[security]) if extracted_data[security].isdigit() else 0
                
                if col_idx == 0:
                    with col1:
                        st.metric(security, f"{amount:,} 百万円")
                elif col_idx == 1:
                    with col2:
                        st.metric(security, f"{amount:,} 百万円")
                else:
                    with col3:
                        st.metric(security, f"{amount:,} 百万円")
            
            # データの修正フォーム
            st.subheader("🔧 データの確認・修正")
            
            # セッション状態の初期化
            if 'confirmed_data' not in st.session_state:
                st.session_state.confirmed_data = None
            
            with st.form("data_correction_form"):
                corrected_year_month = st.text_input("年月 (YYYY-MM形式)", value=extracted_data['年月'])
                
                # 各証券項目の入力フィールドを2列で配置
                col_a, col_b = st.columns(2)
                
                corrected_amounts = {}
                
                with col_a:
                    corrected_amounts['国債'] = st.text_input("国債 (百万円)", value=extracted_data['国債'])
                    corrected_amounts['地方債'] = st.text_input("地方債 (百万円)", value=extracted_data['地方債'])
                    corrected_amounts['短期社債'] = st.text_input("短期社債 (百万円)", value=extracted_data['短期社債'])
                    corrected_amounts['社債'] = st.text_input("社債 (百万円)", value=extracted_data['社債'])
                
                with col_b:
                    corrected_amounts['株式'] = st.text_input("株式 (百万円)", value=extracted_data['株式'])
                    corrected_amounts['外国証券'] = st.text_input("外国証券 (百万円)", value=extracted_data['外国証券'])
                    corrected_amounts['その他の証券'] = st.text_input("その他の証券 (百万円)", value=extracted_data['その他の証券'])
                    corrected_amounts['貸出金'] = st.text_input("貸出金 (百万円)", value=extracted_data.get('貸出金', '0'))
                
                if st.form_submit_button("💾 確定"):
                    try:
                        # データの検証
                        final_data = {'年月': corrected_year_month}
                        all_valid = True
                        
                        for security, value in corrected_amounts.items():
                            if not value.isdigit():
                                st.error(f"❌ {security}の値が正しくありません（数値のみ入力してください）")
                                all_valid = False
                            else:
                                final_data[security] = int(value)
                        
                        if all_valid:
                            # セッション状態に保存
                            st.session_state.confirmed_data = final_data
                            st.success("✅ データが確定されました")
                            
                    except ValueError:
                        st.error("❌ 正しい形式で入力してください")
            
            # 確定されたデータがある場合の処理（フォーム外）
            if st.session_state.confirmed_data:
                final_data = st.session_state.confirmed_data
                
                # 確定データを表示
                st.subheader("📋 確定されたデータ")
                
                # 年月を表示
                st.write(f"**年月: {final_data['年月']}**")
                
                display_data = []
                for security in securities_list:
                    display_data.append({
                        '資産種類': security,
                        '金額（百万円）': f"{final_data[security]:,}"
                    })
                
                df_display = pd.DataFrame(display_data)
                st.dataframe(df_display, use_container_width=True)
                
                # データベース保存とCSVダウンロードのボタンを横並びで配置
                col_save, col_download = st.columns(2)
                
                with col_save:
                    if st.button("💾 データベースに保存", type="primary"):
                        success, result = save_to_database(final_data)
                        if success:
                            if result == "updated":
                                st.success("✅ データベースのレコードを更新しました")
                            else:
                                st.success("✅ データベースに新しいレコードを追加しました")
                        else:
                            st.error(f"❌ データベース保存エラー: {result}")
                
                with col_download:
                    # CSV保存のオプション
                    csv_data = pd.DataFrame([final_data])
                    csv_string = csv_data.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 CSVファイルをダウンロード",
                        data=csv_string,
                        file_name=f"securities_data_{final_data['年月']}.csv",
                        mime="text/csv"
                    )
    else:
        st.info("👆 左側のサイドバーからPDFファイルをアップロードしてください")
    
    # データベースの内容を表示
    st.markdown("---")
    st.subheader("📊 データベースの内容")
    
    # データベースを読み込み
    db_df = load_database()
    
    if not db_df.empty:
        # データを年月順にソート
        db_df_sorted = db_df.sort_values('年月', ascending=True)
        
        # データを表示
        st.dataframe(db_df_sorted, use_container_width=True)
        
        # データベース全体のCSVダウンロード
        csv_all_data = db_df_sorted.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 全データをCSVでダウンロード",
            data=csv_all_data,
            file_name=f"securities_database_{datetime.now().strftime('%Y%m%d')}.csv",
            mime="text/csv"
        )
        
        # 統計情報
        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.metric("総レコード数", len(db_df))
        with col_stat2:
            latest_date = db_df['年月'].max() if '年月' in db_df.columns else "不明"
            st.metric("最新データ", latest_date)
            
    else:
        st.info("データベースにデータがありません。PDFファイルを処理してデータを追加してください。")

elif page == "グラフ表示":
    # グラフ表示ページ
    st.subheader("📈 地方銀行財務データ分析グラフ")
    
    # データベースを読み込み
    db_df = load_database()
    
    if not db_df.empty:
        # データを年月順にソート（昇順：時系列用）
        db_df_sorted = db_df.sort_values('年月', ascending=True)
        
        # 期間選択オプション
        st.subheader("📅 表示期間の選択")
        
        # 利用可能な年月のリストを取得
        available_periods = db_df_sorted['年月'].tolist()
        
        # デフォルトは直近5データ分
        default_start_idx = max(0, len(available_periods) - 5)
        default_end_idx = len(available_periods) - 1
        
        # 期間選択UI
        col_period1, col_period2 = st.columns(2)
        
        with col_period1:
            start_period = st.selectbox(
                "開始年月",
                available_periods,
                index=default_start_idx,
                key="start_period"
            )
        
        with col_period2:
            end_period = st.selectbox(
                "終了年月", 
                available_periods,
                index=default_end_idx,
                key="end_period"
            )
        
        # 期間でデータをフィルタ
        filtered_data = db_df_sorted[
            (db_df_sorted['年月'] >= start_period) & 
            (db_df_sorted['年月'] <= end_period)
        ].copy()
        
        if filtered_data.empty:
            st.error("選択された期間にデータがありません。")
        else:
            # 証券の種類リスト
            securities_columns = ['国債', '地方債', '短期社債', '社債', '株式', '外国証券', 'その他の証券', '貸出金']
            
            # 数値型に変換（エラー処理付き）
            for col in securities_columns:
                filtered_data[col] = pd.to_numeric(filtered_data[col], errors='coerce').fillna(0)
            
            # 各種計算用のデータを準備
            # 有価証券合計（貸出金以外）
            securities_only = ['国債', '地方債', '短期社債', '社債', '株式', '外国証券', 'その他の証券']
            filtered_data['有価証券合計'] = filtered_data[securities_only].sum(axis=1)
            
            # 円債合計
            bonds = ['国債', '地方債', '短期社債', '社債']
            filtered_data['円債合計'] = filtered_data[bonds].sum(axis=1)
            
            # 有価証券と貸出金の合計
            filtered_data['有価証券_貸出金合計'] = filtered_data['有価証券合計'] + filtered_data['貸出金']
            
            # リスク性証券
            risk_securities = ['株式', '外国証券', 'その他の証券']
            
            st.markdown("---")
            
            # グラフ1: 円債の有価証券に占める割合
            st.subheader("📊 グラフ1: 円債の有価証券に占める構成比の変動")
            
            fig1 = go.Figure()
            colors1 = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4']
            markers1 = ['circle', 'square', 'diamond', 'triangle-up']
            
            for i, bond_type in enumerate(bonds):
                # 各円債の有価証券に占める割合を計算
                ratio = filtered_data.apply(
                    lambda row: (row[bond_type] / row['有価証券合計'] * 100) if row['有価証券合計'] > 0 else 0,
                    axis=1
                )
                
                # 開始時点からの変化率を計算（開始時点を0とする）
                base_ratio = ratio.iloc[0] if len(ratio) > 0 else 0
                change_rate = ratio - base_ratio
                
                fig1.add_trace(go.Scatter(
                    name=bond_type,
                    x=filtered_data['年月'],
                    y=change_rate,
                    mode='lines+markers',
                    line=dict(color=colors1[i], width=3),
                    marker=dict(size=10, symbol=markers1[i])
                ))
            
            fig1.update_layout(
                title="円債の有価証券に占める構成比の変動（開始時点からの変化）",
                xaxis_title="年月",
                yaxis_title="構成比変動（%ポイント）",
                height=400,
                xaxis=dict(
                    tickmode='array',
                    tickvals=filtered_data['年月'],
                    ticktext=[f"{period[:4]}年{period[5:7]}月" for period in filtered_data['年月']],
                    tickangle=0
                ),
                yaxis=dict(
                    zeroline=True,
                    zerolinewidth=1,
                    zerolinecolor="rgba(128,128,128,0.8)"
                ),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig1, use_container_width=True)
            
            # グラフ2: 円債と貸出金の有価証券と貸出金に占める割合
            st.subheader("📊 グラフ2: 円債と貸出金の構成比の変動")
            
            fig2 = go.Figure()
            colors2 = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#6BCF7F']
            markers2 = ['circle', 'square', 'diamond', 'triangle-up', 'pentagon']
            
            # 円債の各資産（国債、地方債、短期社債、社債）の有価証券と貸出金に占める割合
            bonds_with_loans = ['国債', '地方債', '短期社債', '社債', '貸出金']
            
            for i, asset in enumerate(bonds_with_loans):
                # 各資産の有価証券と貸出金に占める割合を計算
                ratio = filtered_data.apply(
                    lambda row: (row[asset] / row['有価証券_貸出金合計'] * 100) if row['有価証券_貸出金合計'] > 0 else 0,
                    axis=1
                )
                
                # 開始時点からの変化率を計算（開始時点を0とする）
                base_ratio = ratio.iloc[0] if len(ratio) > 0 else 0
                change_rate = ratio - base_ratio
                
                fig2.add_trace(go.Scatter(
                    name=asset,
                    x=filtered_data['年月'],
                    y=change_rate,
                    mode='lines+markers',
                    line=dict(color=colors2[i], width=3),
                    marker=dict(size=10, symbol=markers2[i])
                ))
            
            fig2.update_layout(
                title="円債と貸出金の構成比の変動（開始時点からの変化）",
                xaxis_title="年月",
                yaxis_title="構成比変動（%ポイント）",
                height=400,
                xaxis=dict(
                    tickmode='array',
                    tickvals=filtered_data['年月'],
                    ticktext=[f"{period[:4]}年{period[5:7]}月" for period in filtered_data['年月']],
                    tickangle=0
                ),
                yaxis=dict(
                    zeroline=True,
                    zerolinewidth=1,
                    zerolinecolor="rgba(128,128,128,0.8)"
                ),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig2, use_container_width=True)
            
            # グラフ3: リスク性証券の有価証券に占める割合
            st.subheader("📊 グラフ3: リスク性証券の有価証券に占める構成比の変動")
            
            fig3 = go.Figure()
            colors3 = ['#FF8A80', '#81C784', '#64B5F6']
            markers3 = ['circle', 'square', 'diamond']
            
            for i, risk_security in enumerate(risk_securities):
                # 各リスク性証券の有価証券に占める割合を計算
                ratio = filtered_data.apply(
                    lambda row: (row[risk_security] / row['有価証券合計'] * 100) if row['有価証券合計'] > 0 else 0,
                    axis=1
                )
                
                # 開始時点からの変化率を計算（開始時点を0とする）
                base_ratio = ratio.iloc[0] if len(ratio) > 0 else 0
                change_rate = ratio - base_ratio
                
                fig3.add_trace(go.Scatter(
                    name=risk_security,
                    x=filtered_data['年月'],
                    y=change_rate,
                    mode='lines+markers',
                    line=dict(color=colors3[i], width=3),
                    marker=dict(size=10, symbol=markers3[i])
                ))
            
            fig3.update_layout(
                title="リスク性証券の有価証券に占める構成比の変動（開始時点からの変化）",
                xaxis_title="年月",
                yaxis_title="構成比変動（%ポイント）",
                height=400,
                xaxis=dict(
                    tickmode='array',
                    tickvals=filtered_data['年月'],
                    ticktext=[f"{period[:4]}年{period[5:7]}月" for period in filtered_data['年月']],
                    tickangle=0
                ),
                yaxis=dict(
                    zeroline=True,
                    zerolinewidth=1,
                    zerolinecolor="rgba(128,128,128,0.8)"
                ),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig3, use_container_width=True)
            
            # グラフ4: 有価証券の構成比
            st.subheader("📊 グラフ4: 有価証券の構成比")
            
            fig4 = go.Figure()
            colors4 = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FF8A80', '#81C784', '#64B5F6']
            
            for i, security in enumerate(securities_only):
                # 各証券の有価証券に占める割合を計算
                ratio = filtered_data.apply(
                    lambda row: (row[security] / row['有価証券合計'] * 100) if row['有価証券合計'] > 0 else 0,
                    axis=1
                )
                
                fig4.add_trace(go.Bar(
                    name=security,
                    x=filtered_data['年月'],
                    y=ratio,
                    marker_color=colors4[i]
                ))
            
            fig4.update_layout(
                title="有価証券の構成比の推移",
                xaxis_title="年月",
                yaxis_title="構成比（%）",
                barmode='stack',
                height=400,
                yaxis=dict(range=[0, 100]),
                xaxis=dict(
                    tickmode='array',
                    tickvals=filtered_data['年月'],
                    ticktext=[f"{period[:4]}年{period[5:7]}月" for period in filtered_data['年月']],
                    tickangle=0
                ),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig4, use_container_width=True)
            
            
            # データテーブル表示
            st.markdown("---")
            st.subheader("📊 選択期間のデータテーブル")
            
            # 表示用データの準備
            display_data = filtered_data[['年月'] + securities_columns].copy()
            
            # 数値を見やすい形式でフォーマット
            for col in securities_columns:
                display_data[col] = display_data[col].apply(lambda x: f"{x:,.0f}")
            
            st.dataframe(display_data, use_container_width=True)
            
            # 統計情報
            st.subheader("📈 統計情報")
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            
            with col_stat1:
                st.metric("選択期間数", len(filtered_data))
            
            with col_stat2:
                latest_securities_total = filtered_data['有価証券合計'].iloc[-1] if len(filtered_data) > 0 else 0
                st.metric("最新期有価証券合計", f"{latest_securities_total:,.0f} 百万円")
            
            with col_stat3:
                latest_loans = filtered_data['貸出金'].iloc[-1] if len(filtered_data) > 0 else 0
                st.metric("最新期貸出金", f"{latest_loans:,.0f} 百万円")
        
    else:
        st.info("📝 データベースにデータがありません。まずはデータ抽出ページでPDFファイルを処理してデータを追加してください。")

# フッター
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: #666; font-size: 0.8em;'>
    地方銀行財務データ抽出システム | 地方銀行主要勘定PDFから貸出金・証券データを抽出します
    </div>
    """,
    unsafe_allow_html=True
)
